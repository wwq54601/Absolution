# backend/services/excel_content_service.py
# Advanced Excel Content Extraction Service - Phase 2A.2
# Provides comprehensive Excel processing with pandas and optional openpyxl

import logging
import os
from typing import Optional, Dict, Any, List, Union
from pathlib import Path
from dataclasses import dataclass

logger = logging.getLogger(__name__)

# Import pandas (required)
try:
    import pandas as pd
    pandas_available = True
except ImportError:
    logger.error("pandas not available - Excel processing will be disabled")
    pandas_available = False

# Import openpyxl (optional, for advanced features)
try:
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    openpyxl_available = True
    advanced_features = True
except ImportError:
    logger.info("openpyxl not available - basic Excel processing only")
    openpyxl_available = False
    advanced_features = False

# Import xlrd (optional, for legacy .xls files)
try:
    import xlrd
    xlrd_available = True
except ImportError:
    logger.info("xlrd not available - .xls files will use pandas engine")
    xlrd_available = False

@dataclass
class WorksheetInfo:
    """Information about a single worksheet"""
    name: str
    index: int
    row_count: int
    column_count: int
    has_data: bool
    visible: bool = True
    
@dataclass
class CellInfo:
    """Information about a cell with value and formatting"""
    value: Any
    formula: Optional[str] = None
    data_type: str = "general"
    address: str = ""
    
@dataclass
class ExcelMetadata:
    """Comprehensive Excel file metadata"""
    worksheets: List[WorksheetInfo]
    total_sheets: int
    file_format: str  # xlsx, xls, xlsm
    has_formulas: bool
    has_charts: bool
    has_images: bool
    workbook_properties: Dict[str, Any]
    cell_count: int
    total_rows: int
    total_columns: int


class ExcelContentExtractor:
    """Advanced Excel content extraction with multiple processing modes."""
    
    def __init__(self):
        self.supported_formats = {'.xlsx', '.xls', '.xlsm', '.xlsb'}
        self.max_file_size_mb = 100  # 100MB limit for Excel files
        self.max_sheets_to_process = 50  # Prevent processing massive workbooks
        self.service_available = pandas_available
        self.advanced_mode = openpyxl_available
        
    def is_excel_file(self, file_path: str) -> bool:
        """Check if file is a supported Excel format."""
        return Path(file_path).suffix.lower() in self.supported_formats
    
    def _validate_file(self, file_path: str) -> Dict[str, Any]:
        """Validate Excel file before processing."""
        validation = {
            'valid': False,
            'error': None,
            'file_size_mb': 0,
            'format': None
        }
        
        if not os.path.exists(file_path):
            validation['error'] = f"File not found: {file_path}"
            return validation
            
        if not self.is_excel_file(file_path):
            validation['error'] = f"Unsupported format: {Path(file_path).suffix}"
            return validation
            
        # Check file size
        file_size_bytes = os.path.getsize(file_path)
        file_size_mb = file_size_bytes / (1024 * 1024)
        
        if file_size_mb > self.max_file_size_mb:
            validation['error'] = f"File too large: {file_size_mb:.1f}MB > {self.max_file_size_mb}MB"
            return validation
            
        validation.update({
            'valid': True,
            'file_size_mb': file_size_mb,
            'format': Path(file_path).suffix.lower().replace('.', '')
        })
        
        return validation
    
    def _extract_workbook_properties(self, file_path: str) -> Dict[str, Any]:
        """Extract workbook-level properties using openpyxl if available."""
        properties = {
            'title': None,
            'author': None,
            'created': None,
            'modified': None,
            'last_saved_by': None,
            'application': None,
            'security': 0
        }
        
        if not openpyxl_available:
            return properties
            
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=False)
            props = workbook.properties
            
            properties.update({
                'title': props.title,
                'author': props.creator,
                'created': str(props.created) if props.created else None,
                'modified': str(props.modified) if props.modified else None,
                'last_saved_by': props.lastModifiedBy,
                'application': props.application,
                'security': getattr(props, 'security', 0)
            })
            
            workbook.close()
            
        except Exception as e:
            logger.warning(f"Failed to extract workbook properties: {e}")
            
        return properties
    
    def _get_worksheet_info(self, file_path: str) -> List[WorksheetInfo]:
        """Get information about all worksheets."""
        worksheets = []
        
        try:
            if openpyxl_available:
                # Use openpyxl for detailed worksheet info
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                for idx, sheet_name in enumerate(workbook.sheetnames):
                    sheet = workbook[sheet_name]
                    
                    # Count non-empty cells
                    row_count = sheet.max_row if sheet.max_row else 0
                    col_count = sheet.max_column if sheet.max_column else 0
                    
                    # Check if sheet has actual data (not just formatting)
                    has_data = False
                    if row_count > 0 and col_count > 0:
                        for row in sheet.iter_rows(max_row=min(10, row_count), max_col=min(10, col_count)):
                            if any(cell.value is not None for cell in row):
                                has_data = True
                                break
                    
                    worksheet_info = WorksheetInfo(
                        name=sheet_name,
                        index=idx,
                        row_count=row_count,
                        column_count=col_count,
                        has_data=has_data,
                        visible=sheet.sheet_state == 'visible'
                    )
                    worksheets.append(worksheet_info)
                
                workbook.close()
                
            else:
                # Fallback to pandas for basic info
                excel_file = pd.ExcelFile(file_path)
                
                for idx, sheet_name in enumerate(excel_file.sheet_names):
                    try:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=0)  # Just headers
                        col_count = len(df.columns)
                        
                        # Read a small sample to estimate rows
                        df_sample = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=1000)
                        row_count = len(df_sample)
                        has_data = not df_sample.empty
                        
                        worksheet_info = WorksheetInfo(
                            name=sheet_name,
                            index=idx,
                            row_count=row_count,
                            column_count=col_count,
                            has_data=has_data
                        )
                        worksheets.append(worksheet_info)
                        
                    except Exception as e:
                        logger.warning(f"Failed to analyze worksheet {sheet_name}: {e}")
                        
                excel_file.close()
                
        except Exception as e:
            logger.error(f"Failed to get worksheet info: {e}")
            
        return worksheets
    
    def _extract_sheet_content(self, file_path: str, sheet_name: str, max_rows: int = 10000) -> Dict[str, Any]:
        """Extract content from a specific worksheet."""
        content = {
            'sheet_name': sheet_name,
            'data': None,
            'text_content': '',
            'formulas': [],
            'data_types': {},
            'has_merged_cells': False,
            'error': None
        }
        
        try:
            # Read data with pandas
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows)
            
            if df.empty:
                content['text_content'] = f"Worksheet '{sheet_name}' is empty"
                return content
            
            # Convert to structured data
            content['data'] = {
                'columns': df.columns.tolist(),
                'rows': df.values.tolist(),
                'shape': df.shape,
                'dtypes': df.dtypes.to_dict()
            }
            
            # Create text representation
            text_lines = [f"Worksheet: {sheet_name}"]
            text_lines.append(f"Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
            text_lines.append("")
            
            # Add column headers
            text_lines.append("Columns: " + ", ".join(str(col) for col in df.columns))
            text_lines.append("")
            
            # Add sample data (first 20 rows)
            sample_rows = min(20, len(df))
            for idx, row in df.head(sample_rows).iterrows():
                row_text = " | ".join(str(val) if pd.notna(val) else "" for val in row.values)
                text_lines.append(f"Row {idx + 1}: {row_text}")
            
            if len(df) > sample_rows:
                text_lines.append(f"... and {len(df) - sample_rows} more rows")
            
            content['text_content'] = "\n".join(text_lines)
            
            # Extract formulas if openpyxl is available
            if openpyxl_available:
                try:
                    workbook = openpyxl.load_workbook(file_path, data_only=False)
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        
                        formulas = []
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.data_type == 'f' and cell.value:  # Formula cell
                                    formulas.append({
                                        'address': cell.coordinate,
                                        'formula': cell.value,
                                        'result': cell.displayed_value
                                    })
                        
                        content['formulas'] = formulas
                        
                        # Check for merged cells
                        content['has_merged_cells'] = len(sheet.merged_cells.ranges) > 0
                    
                    workbook.close()
                    
                except Exception as e:
                    logger.warning(f"Failed to extract formulas from {sheet_name}: {e}")
            
        except Exception as e:
            logger.error(f"Failed to extract content from sheet {sheet_name}: {e}")
            content['error'] = str(e)
            
        return content
    
    def extract_excel_content(self, file_path: str) -> Dict[str, Any]:
        """
        Extract comprehensive content from Excel file.
        
        Returns:
            Dict with keys: 'success', 'metadata', 'worksheets', 'text_content', 'error'
        """
        result = {
            'success': False,
            'metadata': None,
            'worksheets': [],
            'text_content': '',
            'structured_data': {},
            'error': None,
            'processing_info': {
                'pandas_used': pandas_available,
                'openpyxl_used': openpyxl_available,
                'advanced_features': advanced_features
            }
        }
        
        if not self.service_available:
            result['error'] = "Excel processing service not available - pandas required"
            return result
        
        # Validate file
        validation = self._validate_file(file_path)
        if not validation['valid']:
            result['error'] = validation['error']
            return result
        
        logger.info(f"Processing Excel file: {file_path} ({validation['file_size_mb']:.1f}MB)")
        
        try:
            # Extract worksheet information
            worksheets_info = self._get_worksheet_info(file_path)
            
            if not worksheets_info:
                result['error'] = "No worksheets found in Excel file"
                return result
            
            # Limit processing to prevent performance issues
            sheets_to_process = worksheets_info[:self.max_sheets_to_process]
            
            # Extract content from each worksheet
            worksheets_content = []
            all_text_content = []
            
            for sheet_info in sheets_to_process:
                if not sheet_info.has_data:
                    continue
                    
                sheet_content = self._extract_sheet_content(file_path, sheet_info.name)
                worksheets_content.append(sheet_content)
                
                if sheet_content['text_content']:
                    all_text_content.append(sheet_content['text_content'])
            
            # Extract workbook properties
            workbook_properties = self._extract_workbook_properties(file_path)
            
            # Create comprehensive metadata
            total_rows = sum(ws.row_count for ws in worksheets_info)
            total_columns = sum(ws.column_count for ws in worksheets_info)
            has_formulas = any(sheet.get('formulas') for sheet in worksheets_content)
            
            metadata = ExcelMetadata(
                worksheets=worksheets_info,
                total_sheets=len(worksheets_info),
                file_format=validation['format'],
                has_formulas=has_formulas,
                has_charts=False,  # Would need additional analysis
                has_images=False,  # Would need additional analysis
                workbook_properties=workbook_properties,
                cell_count=total_rows * total_columns if worksheets_info else 0,
                total_rows=total_rows,
                total_columns=total_columns
            )
            
            # Combine all text content
            combined_text = f"Excel File: {Path(file_path).name}\n"
            combined_text += f"Format: {validation['format'].upper()}\n"
            combined_text += f"Worksheets: {len(worksheets_info)}\n"
            combined_text += f"Total Data: {total_rows} rows across {len(worksheets_info)} sheets\n\n"
            combined_text += "\n\n".join(all_text_content)
            
            result.update({
                'success': True,
                'metadata': metadata,
                'worksheets': worksheets_content,
                'text_content': combined_text,
                'structured_data': {
                    'workbook_info': {
                        'sheets': [ws.name for ws in worksheets_info],
                        'total_sheets': len(worksheets_info),
                        'format': validation['format']
                    },
                    'worksheets_data': {sheet['sheet_name']: sheet['data'] for sheet in worksheets_content if sheet['data']}
                }
            })
            
            logger.info(f"Successfully processed Excel file: {len(worksheets_content)} sheets, {len(combined_text)} characters")
            
        except Exception as e:
            logger.error(f"Error processing Excel file {file_path}: {e}")
            result['error'] = f"Excel processing failed: {str(e)}"
            
        return result
    
    def get_service_status(self) -> Dict[str, Any]:
        """Get status of Excel content extraction service."""
        return {
            'service_available': self.service_available,
            'pandas_available': pandas_available,
            'openpyxl_available': openpyxl_available,
            'xlrd_available': xlrd_available,
            'advanced_features': advanced_features,
            'supported_formats': list(self.supported_formats),
            'max_file_size_mb': self.max_file_size_mb,
            'max_sheets_to_process': self.max_sheets_to_process
        }


# Singleton instance for global use
excel_extractor = ExcelContentExtractor()


def extract_excel_content(file_path: str) -> Dict[str, Any]:
    """Convenience function for extracting content from Excel file."""
    return excel_extractor.extract_excel_content(file_path)


def is_excel_file(file_path: str) -> bool:
    """Convenience function to check if file is a supported Excel format."""
    return excel_extractor.is_excel_file(file_path)


def get_excel_service_status() -> Dict[str, Any]:
    """Convenience function to get service status."""
    return excel_extractor.get_service_status() 